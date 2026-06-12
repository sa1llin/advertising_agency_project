import os
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

DESKTOP_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(DESKTOP_ROOT))

from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QImage
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QLabel,
    QTableView,
)

from dialogs.order_form_dialog import OrderFormDialog
from pages.all_orders_page import AllOrdersPage
from pages.clients_page import ClientsPage
from pages.new_orders_page import NewOrdersPage
from pages.reports_page import ReportsPage
from pages.analytics_page import AnalyticsPage
from pages.logs_page import LogsPage
from pages.users_page import UsersPage
from permissions import get_nav_items
from models.session import UserSession
from services.report_export import (
    build_report_html,
    export_report_csv,
    export_report_xlsx,
)
from services.invoice_service import (
    amount_in_words,
    build_invoice_lines,
    generate_invoice,
    invoice_number,
)
from windows.main_window import MainWindow


class FakeApiClient:
    pass


CLIENTS = [
    {
        "id": 1,
        "client_type": "company",
        "full_name": "Олена Коваль",
        "company_name": "ТОВ Приклад",
        "phone": "+380501112233",
        "email": "office@example.ua",
        "legal_address": "Київ",
        "tax_number": "12345678",
        "comment": None,
        "is_active": True,
        "created_at": "2026-06-10T09:00:00",
        "updated_at": "2026-06-10T09:00:00",
    }
]

ORDERS = [
    {
        "id": 10,
        "order_number": "ORD-00010",
        "client_id": 1,
        "manager_id": None,
        "order_type": "billboard",
        "status": "new",
        "order_date": "2026-06-10T09:30:00",
        "rental_start": "2026-06-12",
        "rental_end": "2026-06-20",
        "total_amount": "12000.00",
        "vat_amount": "2000.00",
        "discount_amount": "0.00",
    }
]

APPLICATIONS = [
    {
        "id": 20,
        "full_name": "Олена Коваль",
        "phone": "+380501112233",
        "email": "olena@example.com",
        "service_type": "billboard",
        "comment": "Потрібна консультація",
        "source": "calculator",
        "calculation_data": {
            "service_type": "billboard",
            "advertising_space_id": 10,
            "location": "Центральна площа, 1",
            "size": "3x6",
            "period_start": "2026-06-12",
            "period_end": "2026-06-20",
            "days": 9,
            "need_printing": True,
            "estimated_total": "12000.00",
            "price_rows": [],
        },
        "estimated_total": "12000.00",
        "status": "new",
        "client_id": None,
        "order_id": None,
        "processed_by": None,
        "is_hidden": False,
        "submitted_at": "2026-06-10T09:30:00",
        "processed_at": None,
    }
]

CATALOG = {
    "advertising_spaces": [
        {
            "id": 10,
            "title": "Білборд",
            "space_type": "billboard",
            "location": "Центральна площа, 1",
            "size": "3x6",
            "base_price": "850.00",
            "is_active": True,
        },
        {
            "id": 20,
            "title": "LED",
            "space_type": "led",
            "location": "Соборна, 12",
            "size": "6x3",
            "base_price": "1.20",
            "is_active": True,
        },
    ],
    "pricing_items": [
        {
            "category": "billboard_print",
            "code": "3x6",
            "label": "Плакат 3x6",
            "amount": "1800.00",
            "is_active": True,
        },
        {
            "category": "print_product",
            "code": "business_card",
            "label": "Візитки",
            "amount": "2.50",
            "is_active": True,
        },
        {
            "category": "print_product",
            "code": "other",
            "label": "Інше",
            "amount": "1.00",
            "is_active": True,
        },
        {
            "category": "print_material",
            "code": "coated_paper",
            "label": "Папір",
            "amount": "0.50",
            "is_active": True,
        },
        {
            "category": "print_size",
            "code": "small",
            "label": "Малий",
            "amount": "0.00",
            "is_active": True,
        },
        {
            "category": "print_color",
            "code": "full_color",
            "label": "Повноколірна",
            "amount": "1.50",
            "is_active": True,
        },
    ],
}


class PagePopulationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def test_new_orders_page_populates_cards_and_table(self):
        page = NewOrdersPage(FakeApiClient())
        page._apply_data(
            {
                "applications": APPLICATIONS,
                "clients": CLIENTS,
                "managers": [],
                "catalog": CATALOG,
            }
        )

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.total_card.value_label.text(), "1")
        self.assertEqual(page.table.item(0, 0).text(), "Олена Коваль")
        self.assertEqual(page.table.item(0, 3).text(), "Білборд")

    def test_regular_application_changes_only_total_counter(self):
        page = NewOrdersPage(FakeApiClient())
        page._apply_data(
            {
                "applications": [
                    {
                        **APPLICATIONS[0],
                        "id": 21,
                        "service_type": "other",
                        "source": "contact",
                        "calculation_data": None,
                        "estimated_total": None,
                    }
                ],
                "clients": CLIENTS,
                "managers": [],
                "catalog": CATALOG,
            }
        )

        self.assertEqual(page.total_card.value_label.text(), "1")
        self.assertEqual(page.billboard_card.value_label.text(), "0")
        self.assertEqual(page.led_card.value_label.text(), "0")
        self.assertEqual(page.printing_card.value_label.text(), "0")

    def test_calculator_application_prefills_order_form(self):
        page = NewOrdersPage(FakeApiClient(), "manager")
        dialog = OrderFormDialog(CLIENTS, [], "manager", CATALOG)
        application = {
            **APPLICATIONS[0],
            "client_id": 1,
        }

        page._prefill_order_dialog(dialog, application)

        editor = dialog.segment_editors[0]
        self.assertEqual(dialog.client_combo.currentData(), 1)
        self.assertEqual(dialog.order_type.currentData(), "billboard")
        self.assertEqual(editor.space_combo.currentData(), 10)
        self.assertEqual(
            editor.period_start.date().toString("yyyy-MM-dd"),
            "2026-06-12",
        )
        self.assertEqual(
            editor.period_end.date().toString("yyyy-MM-dd"),
            "2026-06-20",
        )
        self.assertTrue(editor.need_printing.isChecked())
        self.assertIn("12000.00 грн", dialog.comment.toPlainText())

    def test_all_orders_page_populates_table(self):
        page = AllOrdersPage(FakeApiClient())
        page._apply_data(
            {
                "orders": ORDERS,
                "clients": CLIENTS,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
                "catalog": CATALOG,
            }
        )

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 0).text(), "ORD-00010")
        self.assertEqual(page.table.item(0, 3).text(), "ТОВ Приклад")
        self.assertFalse(page.invoice_button.isEnabled())
        page.table.selectRow(0)
        self.app.processEvents()
        self.assertTrue(page.invoice_button.isEnabled())

    def test_invoice_generation_uses_service_template_and_order_data(self):
        order = {
            **ORDERS[0],
            "vat_rate": "20.00",
            "discount_rate": "5.00",
            "amount_without_vat": "10000.00",
            "discount_amount": "500.00",
            "vat_amount": "1900.00",
            "total_amount": "11400.00",
            "segments": [
                {
                    "advertising_space_id": 10,
                    "period_start": "2026-06-12",
                    "period_end": "2026-06-20",
                    "need_printing": True,
                    "rental_cost": "8200.00",
                    "printing_cost": "1800.00",
                    "placement_cost": "0.00",
                    "materials_cost": "0.00",
                    "subtotal": "10000.00",
                }
            ],
        }
        lines = build_invoice_lines(order, CATALOG)

        self.assertEqual(len(lines), 1)
        self.assertIn("Центральна площа, 1", lines[0].description)
        self.assertIn("друком плаката", lines[0].description)
        self.assertEqual(invoice_number(order, date(2026, 6, 12)), "RF-2026-00010")
        self.assertEqual(
            amount_in_words("11400.00"),
            "Одинадцять тисяч чотириста гривень 00 копійок",
        )

        with tempfile.TemporaryDirectory() as directory:
            png_path = Path(directory) / "invoice.png"
            pdf_path = Path(directory) / "invoice.pdf"
            for path in (png_path, pdf_path):
                generate_invoice(
                    path,
                    order,
                    CLIENTS[0],
                    "Менеджер Один",
                    CATALOG,
                    invoice_date=date(2026, 6, 12),
                )
                self.assertTrue(path.exists())
                self.assertGreater(path.stat().st_size, 10_000)

            image = QImage(str(png_path))
            self.assertEqual((image.width(), image.height()), (1055, 1491))

            variants = [
                (
                    {
                        **order,
                        "order_type": "led",
                        "segments": [
                            {
                                "advertising_space_id": 20,
                                "period_start": "2026-06-12",
                                "period_end": "2026-06-20",
                                "video_seconds": 10,
                                "impressions_per_day": 100,
                                "subtotal": "10000.00",
                            }
                        ],
                    },
                    Path(directory) / "invoice_led.png",
                    (1055, 1491),
                ),
                (
                    {
                        **order,
                        "order_type": "printing",
                        "rental_start": None,
                        "rental_end": None,
                        "segments": [
                            {
                                "product_type": "business_card",
                                "product_name": "Візитки",
                                "material_code": "coated_paper",
                                "size_code": "small",
                                "color_mode": "full_color",
                                "quantity": 500,
                                "subtotal": "10000.00",
                            }
                        ],
                    },
                    Path(directory) / "invoice_printing.png",
                    (1055, 1491),
                ),
            ]
            for variant, path, expected_size in variants:
                generate_invoice(
                    path,
                    variant,
                    CLIENTS[0],
                    "Менеджер Один",
                    CATALOG,
                    invoice_date=date(2026, 6, 12),
                )
                rendered = QImage(str(path))
                self.assertEqual(
                    (rendered.width(), rendered.height()),
                    expected_size,
                )

    def test_all_orders_page_creates_invoice_for_selected_order(self):
        page = AllOrdersPage(FakeApiClient())
        page._apply_data(
            {
                "orders": ORDERS,
                "clients": CLIENTS,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
                "catalog": CATALOG,
            }
        )
        page.table.selectRow(0)
        self.app.processEvents()

        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "invoice.pdf"
            with (
                patch(
                    "pages.all_orders_page.QFileDialog.getSaveFileName",
                    return_value=(str(destination), "PDF (*.pdf)"),
                ),
                patch(
                    "pages.all_orders_page.generate_invoice",
                    return_value=destination,
                ) as generate,
                patch("pages.all_orders_page.QMessageBox.information"),
            ):
                page.create_invoice()

            generate.assert_called_once_with(
                destination,
                ORDERS[0],
                CLIENTS[0],
                "Не призначено",
                CATALOG,
            )

    def test_crm_table_uses_model_view_and_single_row_selection(self):
        page = AllOrdersPage(FakeApiClient())

        self.assertIsInstance(page.table, QTableView)
        self.assertEqual(
            page.table.selectionMode(),
            QAbstractItemView.SelectionMode.SingleSelection,
        )
        self.assertTrue(page.table.isSortingEnabled())
        self.assertEqual(
            page.table.horizontalScrollBarPolicy(),
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn,
        )

    def test_sidebar_starts_with_user_account_without_brand_logo(self):
        session = UserSession(
            user_id=1,
            login="admin",
            full_name="Адміністратор",
            position="Адміністратор",
            role="admin",
        )
        with patch.object(MainWindow, "set_active_page"):
            window = MainWindow(session, FakeApiClient())

        self.assertIsNone(window.findChild(QLabel, "brandMark"))
        self.assertIsNotNone(window.findChild(QLabel, "userName"))
        window.close()

    def test_all_orders_page_combines_search_and_filters(self):
        page = AllOrdersPage(FakeApiClient())
        orders = [
            {**ORDERS[0], "manager_id": 5},
            {
                **ORDERS[0],
                "id": 11,
                "order_number": "ORD-00011",
                "client_id": 2,
                "manager_id": 6,
                "order_type": "printing",
                "status": "completed",
                "order_date": "2026-06-11T11:00:00",
            },
        ]
        clients = CLIENTS + [
            {
                **CLIENTS[0],
                "id": 2,
                "company_name": None,
                "full_name": "Другий клієнт",
            }
        ]
        page._apply_data(
            {
                "orders": orders,
                "clients": clients,
                "client_names": {
                    1: "ТОВ Приклад",
                    2: "Другий клієнт",
                },
                "managers": [
                    {"id": 5, "username": "one", "full_name": "Менеджер Один"},
                    {"id": 6, "username": "two", "full_name": "Менеджер Два"},
                ],
                "catalog": CATALOG,
            }
        )

        page.search_input.setText("Менеджер Два")
        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 0).text(), "ORD-00011")

        page.search_input.clear()
        page.status_filter.setCurrentIndex(page.status_filter.findData("completed"))
        page.type_filter.setCurrentIndex(page.type_filter.findData("printing"))
        self.assertEqual(page.table.rowCount(), 1)

        page.status_filter.setCurrentIndex(0)
        page.type_filter.setCurrentIndex(0)
        page.date_mode.setCurrentIndex(page.date_mode.findData("date"))
        page.date_from.setDate(QDate(2026, 6, 10))
        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 0).text(), "ORD-00010")

    def test_order_form_builds_role_aware_payload(self):
        manager_dialog = OrderFormDialog(
            CLIENTS,
            [],
            "manager",
            CATALOG,
        )
        manager_dialog.order_type.setCurrentIndex(
            manager_dialog.order_type.findData("printing")
        )
        editor = manager_dialog.segment_editors[0]
        editor.product_type.setCurrentIndex(
            editor.product_type.findData("business_card")
        )
        editor.quantity.setValue(100)
        manager_payload = manager_dialog.get_data()

        admin_dialog = OrderFormDialog(
            CLIENTS,
            [{"id": 5, "username": "manager", "full_name": "Менеджер"}],
            "admin",
            CATALOG,
        )
        admin_dialog.manager_combo.setCurrentIndex(
            admin_dialog.manager_combo.findData(5)
        )
        admin_payload = admin_dialog.get_data()

        self.assertNotIn("manager_id", manager_payload)
        self.assertEqual(
            manager_payload["segments"][0]["product_type"],
            "business_card",
        )
        self.assertEqual(admin_payload["manager_id"], 5)

    def test_order_form_supports_split_period_and_extension(self):
        dialog = OrderFormDialog(CLIENTS, [], "manager", CATALOG)
        first = dialog.segment_editors[0]
        first.space_combo.setCurrentIndex(first.space_combo.findData(10))
        first.period_start.setDate(QDate(2026, 6, 21))
        first.period_end.setDate(QDate(2026, 6, 30))
        dialog.add_segment()
        second = dialog.segment_editors[1]
        second.space_combo.setCurrentIndex(second.space_combo.findData(10))
        second.period_start.setDate(QDate(2026, 7, 1))
        second.period_end.setDate(QDate(2026, 7, 18))

        payload = dialog.get_data()
        self.assertEqual(len(payload["segments"]), 2)
        self.assertEqual(payload["segments"][0]["period_end"], "2026-06-30")
        self.assertEqual(payload["segments"][1]["period_start"], "2026-07-01")

        existing = {
            **ORDERS[0],
            "segments": [
                {
                    "segment_kind": "initial",
                    "advertising_space_id": 10,
                    "period_start": "2026-06-21",
                    "period_end": "2026-07-18",
                    "need_printing": False,
                }
            ],
        }
        prolong = OrderFormDialog(
            CLIENTS,
            [],
            "manager",
            CATALOG,
            order=existing,
            prolong_only=True,
        )
        extension_payload = prolong.get_data()
        self.assertEqual(
            extension_payload["segments"][0]["period_start"],
            "2026-07-19",
        )
        self.assertEqual(
            extension_payload["segments"][0]["segment_kind"],
            "extension",
        )

    def test_clients_page_populates_table(self):
        page = ClientsPage(FakeApiClient(), "admin")
        page._apply_data(CLIENTS)

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 0).text(), "Юридична особа")
        self.assertEqual(page.table.item(0, 1).text(), "ТОВ Приклад")

    def test_clients_page_filters_by_type(self):
        page = ClientsPage(FakeApiClient(), "admin")
        page._apply_data(
            CLIENTS
            + [
                {
                    **CLIENTS[0],
                    "id": 2,
                    "client_type": "individual",
                    "company_name": None,
                    "full_name": "Марія Іваненко",
                    "phone": "+380671111111",
                    "email": "maria@example.ua",
                    "legal_address": None,
                }
            ]
        )

        page.type_filter.setCurrentIndex(page.type_filter.findData("individual"))

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 1).text(), "Марія Іваненко")

    def test_clients_page_filters_by_active_status(self):
        page = ClientsPage(FakeApiClient(), "admin")
        page._apply_data(
            CLIENTS
            + [
                {
                    **CLIENTS[0],
                    "id": 2,
                    "company_name": None,
                    "full_name": "Неактивний клієнт",
                    "is_active": False,
                }
            ]
        )

        page.status_filter.setCurrentIndex(page.status_filter.findData(False))

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 1).text(), "Неактивний клієнт")

    def test_completed_order_cannot_be_prolonged(self):
        page = AllOrdersPage(FakeApiClient())
        page._apply_data(
            {
                "orders": [{**ORDERS[0], "status": "completed"}],
                "clients": CLIENTS,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
                "catalog": CATALOG,
            }
        )

        page.table.selectRow(0)
        self.app.processEvents()

        self.assertFalse(page.prolong_button.isEnabled())

    def test_manager_gets_deactivate_action(self):
        page = ClientsPage(FakeApiClient(), "manager")

        self.assertEqual(page.delete_button.text(), "Деактивувати")

    def test_reports_page_populates_financial_columns(self):
        page = ReportsPage(FakeApiClient())
        page._apply_data(
            {
                "orders": ORDERS,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
            }
        )

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 5).text(), "10 000.00 грн")
        self.assertEqual(page.table.item(0, 6).text(), "2 000.00 грн")
        self.assertEqual(page.table.item(0, 7).text(), "0.00 грн")
        self.assertEqual(page.table.item(0, 8).text(), "12 000.00 грн")
        self.assertIn("12 000.00 грн", page.status_label.text())
        self.assertTrue(page.export_button.isEnabled())
        self.assertTrue(page.print_button.isEnabled())

    def test_reports_page_filters_by_service_and_overlapping_period(self):
        page = ReportsPage(FakeApiClient())
        orders = [
            {
                **ORDERS[0],
                "segments": [
                    {
                        "period_start": "2026-06-28",
                        "period_end": "2026-07-05",
                    }
                ],
            },
            {
                **ORDERS[0],
                "id": 11,
                "order_number": "ORD-00011",
                "order_type": "printing",
                "order_date": "2026-07-15T10:00:00",
                "rental_start": None,
                "rental_end": None,
                "manager_id": 5,
                "amount_without_vat": "5000.00",
                "vat_amount": "900.00",
                "discount_amount": "500.00",
                "total_amount": "5400.00",
            },
        ]
        page._apply_data(
            {
                "orders": orders,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [
                    {
                        "id": 5,
                        "username": "manager",
                        "full_name": "Менеджер Один",
                    }
                ],
            }
        )

        page.period_filter.setCurrentIndex(page.period_filter.findData("custom"))
        page.date_from.setDate(QDate(2026, 7, 1))
        page.date_to.setDate(QDate(2026, 7, 31))
        page.generate_report()
        self.assertEqual(page.table.rowCount(), 2)

        page.type_filter.setCurrentIndex(page.type_filter.findData("printing"))
        page.generate_report()
        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 0).text(), "ORD-00011")
        self.assertEqual(page.table.item(0, 4).text(), "Менеджер Один")

    def test_reports_page_lists_client_orders_separately(self):
        page = ReportsPage(FakeApiClient())
        orders = [
            {
                **ORDERS[0],
                "amount_without_vat": "10000.00",
            },
            {
                **ORDERS[0],
                "id": 11,
                "order_number": "ORD-00011",
                "order_type": "printing",
                "order_date": "2026-07-15T10:00:00",
                "rental_start": None,
                "rental_end": None,
                "amount_without_vat": "5000.00",
                "vat_amount": "900.00",
                "discount_amount": "500.00",
                "total_amount": "5400.00",
                "segments": [
                    {
                        "product_type": "business_card",
                        "product_name": "Візитки",
                        "quantity": 500,
                        "size_code": "90x50",
                        "material_code": "крейдований папір",
                    }
                ],
            },
            {
                **ORDERS[0],
                "id": 12,
                "order_number": "ORD-00012",
                "client_id": 2,
                "order_type": "printing",
                "product_name": "Чашки",
                "quantity": 20,
                "order_date": "2026-07-20T10:00:00",
                "rental_start": None,
                "rental_end": None,
                "amount_without_vat": "2000.00",
                "vat_amount": "400.00",
                "discount_amount": "0.00",
                "total_amount": "2400.00",
            },
        ]
        page._apply_data(
            {
                "orders": orders,
                "client_names": {
                    1: "ТОВ Приклад",
                    2: "ФОП Інший клієнт",
                },
                "managers": [],
            }
        )

        page.report_mode.setCurrentIndex(page.report_mode.findData("clients"))
        self.assertEqual(page.table.rowCount(), 3)
        page.entity_filter.setCurrentIndex(page.entity_filter.findData(1))

        self.assertEqual(page.table.rowCount(), 2)
        self.assertEqual(page.table.item(0, 0).text(), "ТОВ Приклад")
        self.assertEqual(page.table.item(0, 1).text(), "ORD-00010")
        self.assertEqual(page.table.item(1, 1).text(), "ORD-00011")
        self.assertIn("Візитки", page.table.item(1, 2).text())
        self.assertIn("500 шт.", page.table.item(1, 2).text())
        self.assertNotIn(
            "Чашки",
            " ".join(page.table.item(row, 2).text() for row in range(2)),
        )
        self.assertEqual(page.table.item(0, 6).text(), "10 000.00 грн")
        self.assertEqual(page.table.item(1, 9).text(), "5 400.00 грн")
        self.assertEqual(page.totals["order_count"], 2)
        self.assertEqual(page.totals["total_amount"], 17400)

        page.period_filter.setCurrentIndex(page.period_filter.findData("custom"))
        page.date_from.setDate(QDate(2026, 7, 1))
        page.date_to.setDate(QDate(2026, 7, 31))
        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 1).text(), "ORD-00011")
        self.assertEqual(page.totals["order_count"], 1)

        page._show_error("test")
        self.assertEqual(page.table.rowCount(), 0)
        self.assertEqual(page.totals["order_count"], 0)
        self.assertIn("Помилка завантаження", page.status_label.text())

    def test_reports_page_calculates_billboard_days_without_utilization(self):
        page = ReportsPage(FakeApiClient())
        catalog = {
            **CATALOG,
            "advertising_spaces": [
                *CATALOG["advertising_spaces"],
                {
                    **CATALOG["advertising_spaces"][0],
                    "id": 11,
                    "location": "Проспект Науки, 10",
                },
            ],
        }
        orders = [
            {
                **ORDERS[0],
                "order_number": "ORD-SPACE-1",
                "amount_without_vat": "800.00",
                "discount_amount": "80.00",
                "vat_amount": "144.00",
                "total_amount": "864.00",
                "segments": [
                    {
                        "advertising_space_id": 10,
                        "period_start": "2026-06-28",
                        "period_end": "2026-07-05",
                        "rental_cost": "800.00",
                        "placement_cost": "0.00",
                        "subtotal": "800.00",
                    }
                ],
            },
            {
                **ORDERS[0],
                "id": 11,
                "order_number": "ORD-SPACE-2",
                "rental_start": "2026-07-04",
                "rental_end": "2026-07-10",
                "amount_without_vat": "700.00",
                "discount_amount": "0.00",
                "vat_amount": "140.00",
                "total_amount": "840.00",
                "segments": [
                    {
                        "advertising_space_id": 10,
                        "period_start": "2026-07-04",
                        "period_end": "2026-07-10",
                        "rental_cost": "700.00",
                        "placement_cost": "0.00",
                        "subtotal": "700.00",
                    }
                ],
            },
            {
                **ORDERS[0],
                "id": 12,
                "order_number": "ORD-OTHER-SPACE",
                "rental_start": "2026-07-11",
                "rental_end": "2026-07-12",
                "amount_without_vat": "200.00",
                "discount_amount": "0.00",
                "vat_amount": "40.00",
                "total_amount": "240.00",
                "segments": [
                    {
                        "advertising_space_id": 11,
                        "period_start": "2026-07-11",
                        "period_end": "2026-07-12",
                        "rental_cost": "200.00",
                        "placement_cost": "0.00",
                        "subtotal": "200.00",
                    }
                ],
            },
        ]
        page._apply_data(
            {
                "orders": orders,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
                "catalog": catalog,
            }
        )

        page.report_mode.setCurrentIndex(page.report_mode.findData("spaces"))
        page.period_filter.setCurrentIndex(page.period_filter.findData("custom"))
        page.date_from.setDate(QDate(2026, 7, 1))
        page.date_to.setDate(QDate(2026, 7, 31))
        self.assertEqual(page.table.rowCount(), 2)
        page.entity_filter.setCurrentIndex(page.entity_filter.findData(10))

        self.assertEqual(page.table.rowCount(), 1)
        self.assertIn("Центральна площа, 1", page.table.item(0, 0).text())
        self.assertEqual(page.table.item(0, 3).text(), "2")
        self.assertEqual(page.table.item(0, 5).text(), "10")
        self.assertEqual(page.table.item(0, 6).text(), "—")
        self.assertEqual(page.table.item(0, 7).text(), "1 150.00 грн")
        self.assertIn("ORD-SPACE-1", page.table.item(0, 4).text())
        self.assertIn("ORD-SPACE-2", page.table.item(0, 4).text())
        self.assertNotIn("ORD-OTHER-SPACE", page.table.item(0, 4).text())

        with tempfile.TemporaryDirectory() as directory:
            xlsx_path = Path(directory) / "spaces.xlsx"
            export_report_xlsx(
                str(xlsx_path),
                page._period_label(),
                page.report_rows,
                page.totals,
                report_title=page._report_title(),
                columns=page.current_columns,
            )

            from openpyxl import load_workbook

            workbook = load_workbook(xlsx_path, data_only=False)
            sheet = workbook["Звіт по рекламних площинах"]
            self.assertEqual(sheet["F6"].value, 10)
            self.assertEqual(sheet["H6"].value, 1150)
            self.assertEqual(sheet["H7"].value, 1150)
            workbook.close()

    def test_reports_page_calculates_led_utilization_from_reserved_seconds(self):
        page = ReportsPage(FakeApiClient())
        led_orders = []
        reservations = [
            ("ORD-LED-A", 30, 100, "2026-01-15"),
            ("ORD-LED-B", 10, 300, "2026-01-25"),
            ("ORD-LED-C", 20, 450, "2026-01-12"),
            ("ORD-LED-D", 30, 1000, "2026-01-13"),
            ("ORD-LED-E", 20, 920, "2026-01-19"),
        ]
        for index, (number, seconds, impressions, period_end) in enumerate(
            reservations,
            start=1,
        ):
            daily_seconds = seconds * impressions
            led_orders.append(
                {
                    **ORDERS[0],
                    "id": 100 + index,
                    "order_number": number,
                    "order_type": "led",
                    "status": "new",
                    "segments": [
                        {
                            "advertising_space_id": 20,
                            "period_start": "2026-01-01",
                            "period_end": period_end,
                            "video_seconds": seconds,
                            "impressions_per_day": impressions,
                            "rental_cost": "0.00",
                            "placement_cost": str(daily_seconds),
                            "subtotal": str(daily_seconds),
                        }
                    ],
                }
            )
        led_orders.append(
            {
                **ORDERS[0],
                "id": 200,
                "order_number": "ORD-LED-CANCELLED",
                "order_type": "led",
                "status": "cancelled",
                "segments": [
                    {
                        "advertising_space_id": 20,
                        "period_start": "2026-01-01",
                        "period_end": "2026-01-12",
                        "video_seconds": 20,
                        "impressions_per_day": 70,
                        "rental_cost": "0.00",
                        "placement_cost": "1400.00",
                        "subtotal": "1400.00",
                    }
                ],
            }
        )
        page._apply_data(
            {
                "orders": led_orders,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
                "catalog": CATALOG,
            }
        )

        page.report_mode.setCurrentIndex(page.report_mode.findData("spaces"))
        page.period_filter.setCurrentIndex(page.period_filter.findData("custom"))
        page.date_from.setDate(QDate(2026, 1, 1))
        page.date_to.setDate(QDate(2026, 1, 12))

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 3).text(), "5")
        self.assertEqual(page.table.item(0, 5).text(), "12")
        self.assertEqual(page.table.item(0, 6).text(), "97.84 %")
        self.assertNotIn(
            "ORD-LED-CANCELLED",
            page.table.item(0, 4).text(),
        )

    def test_reports_exclude_cancelled_orders_and_their_amounts(self):
        page = ReportsPage(FakeApiClient())
        active = {
            **ORDERS[0],
            "amount_without_vat": "1000.00",
            "vat_amount": "200.00",
            "discount_amount": "0.00",
            "total_amount": "1200.00",
            "segments": [
                {
                    "advertising_space_id": 10,
                    "period_start": "2026-06-12",
                    "period_end": "2026-06-20",
                    "rental_cost": "1000.00",
                    "placement_cost": "0.00",
                    "subtotal": "1000.00",
                }
            ],
        }
        cancelled = {
            **active,
            "id": 11,
            "order_number": "ORD-CANCELLED",
            "status": "cancelled",
            "amount_without_vat": "9000.00",
            "vat_amount": "1800.00",
            "total_amount": "10800.00",
            "segments": [
                {
                    "advertising_space_id": 10,
                    "period_start": "2026-06-21",
                    "period_end": "2026-06-30",
                    "rental_cost": "9000.00",
                    "placement_cost": "0.00",
                    "subtotal": "9000.00",
                }
            ],
        }
        page._apply_data(
            {
                "orders": [active, cancelled],
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
                "catalog": CATALOG,
            }
        )

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.totals["total_amount"], 1200)
        self.assertNotIn("ORD-CANCELLED", page.table.item(0, 0).text())

        page.report_mode.setCurrentIndex(page.report_mode.findData("clients"))
        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.totals["order_count"], 1)
        self.assertEqual(page.totals["total_amount"], 1200)

        page.report_mode.setCurrentIndex(page.report_mode.findData("spaces"))
        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 3).text(), "1")
        self.assertEqual(page.table.item(0, 7).text(), "1 000.00 грн")
        self.assertNotIn("ORD-CANCELLED", page.table.item(0, 4).text())

    def test_report_export_and_print_document(self):
        page = ReportsPage(FakeApiClient())
        page._apply_data(
            {
                "orders": ORDERS,
                "client_names": {1: "ТОВ Приклад"},
                "managers": [],
            }
        )

        with tempfile.TemporaryDirectory() as directory:
            xlsx_path = Path(directory) / "report.xlsx"
            csv_path = Path(directory) / "report.csv"
            export_report_xlsx(
                str(xlsx_path),
                page._period_label(),
                page.report_rows,
                page.totals,
            )
            export_report_csv(
                str(csv_path),
                page._period_label(),
                page.report_rows,
                page.totals,
            )

            self.assertTrue(xlsx_path.exists())
            self.assertGreater(xlsx_path.stat().st_size, 1000)
            from openpyxl import load_workbook

            workbook = load_workbook(xlsx_path, data_only=False)
            sheet = workbook["Звіт по замовленнях"]
            self.assertEqual(sheet["A6"].value, "ORD-00010")
            self.assertEqual(sheet["I6"].value, 12000)
            workbook.close()
            self.assertIn(
                "ORD-00010",
                csv_path.read_text(encoding="utf-8-sig"),
            )

        html = build_report_html(
            page._period_label(),
            page.report_rows,
            page.totals,
        )
        self.assertIn("ORD-00010", html)
        self.assertIn("12 000.00 грн", html)

    def test_users_page_populates_employee_table(self):
        page = UsersPage(FakeApiClient(), current_user_id=1)
        page._apply_data(
            [
                {
                    "id": 1,
                    "username": "admin",
                    "full_name": "Адміністратор",
                    "role": "admin",
                    "email": "admin@example.com",
                    "phone": None,
                    "is_active": True,
                    "created_at": "2026-06-10T09:00:00",
                }
            ]
        )

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 2).text(), "Адміністратор")

    def test_logs_page_populates_audit_table(self):
        page = LogsPage(FakeApiClient())
        page._apply_data(
            [
                {
                    "id": 1,
                    "username": "admin",
                    "action": "user_created",
                    "entity_name": "user",
                    "entity_id": 2,
                    "details": None,
                    "created_at": "2026-06-10T09:00:00",
                }
            ]
        )

        self.assertEqual(page.table.rowCount(), 1)
        self.assertEqual(page.table.item(0, 2).text(), "Створено працівника")

    def test_analytics_page_populates_summary(self):
        page = AnalyticsPage(FakeApiClient())
        page._apply_data(
            {
                "orders_total": 5,
                "clients_total": 3,
                "active_managers": 2,
                "total_revenue": "1000.00",
                "orders_by_status": {"new": 1, "in_progress": 2, "completed": 2},
            }
        )

        self.assertEqual(page.orders_card.value_label.text(), "5")
        self.assertEqual(page.table.item(0, 1).text(), "1")

    def test_manager_has_analytics_navigation(self):
        manager_keys = {item["key"] for item in get_nav_items("manager")}

        self.assertIn("analytics", manager_keys)
        self.assertNotIn("users", manager_keys)
        self.assertNotIn("logs", manager_keys)


if __name__ == "__main__":
    unittest.main()
