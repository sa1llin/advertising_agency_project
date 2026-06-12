import csv
from datetime import datetime
from html import escape
from pathlib import Path

ReportColumn = tuple[str, str, str]

REPORT_COLUMNS: list[ReportColumn] = [
    ("order_number", "Номер замовлення", "text"),
    ("client", "Клієнт", "text"),
    ("period", "Період", "text"),
    ("service", "Тип послуги", "text"),
    ("manager", "Менеджер", "text"),
    ("sale_amount", "Сума продажу", "money"),
    ("vat_amount", "ПДВ", "money"),
    ("discount_amount", "Знижка", "money"),
    ("total_amount", "Підсумкова сума", "money"),
]


def export_report_xlsx(
    path: str,
    period_label: str,
    rows: list[dict[str, object]],
    totals: dict[str, object],
    *,
    report_title: str = "Звіт по замовленнях",
    columns: list[ReportColumn] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    selected_columns = columns or REPORT_COLUMNS
    column_count = len(selected_columns)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(report_title)
    sheet.freeze_panes = "A6"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=column_count,
    )
    sheet["A1"] = "Creative Spark Agency CRM"
    sheet["A1"].font = Font(size=18, bold=True, color="071B3A")
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=column_count,
    )
    sheet["A2"] = report_title
    sheet["A2"].font = Font(size=14, bold=True, color="071B3A")
    sheet["A2"].alignment = Alignment(horizontal="center")

    sheet.merge_cells(
        start_row=3,
        start_column=1,
        end_row=3,
        end_column=column_count,
    )
    sheet["A3"] = (
        f"Період: {period_label}. "
        f"Сформовано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    sheet["A3"].font = Font(size=10, color="6F7D95")
    sheet["A3"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="06284A")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="DDE5F0"))
    for column, (_, header, _) in enumerate(selected_columns, start=1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, report_row in enumerate(rows, start=6):
        for column, (key, _, value_type) in enumerate(
            selected_columns,
            start=1,
        ):
            value = report_row.get(key, "")
            if value_type in ("money", "number"):
                value = float(value or 0)
            elif value_type == "integer":
                value = int(value or 0)
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal=(
                    "right" if value_type in ("money", "number", "integer") else "left"
                ),
                vertical="center",
                wrap_text=True,
            )
            if value_type == "money":
                cell.number_format = '#,##0.00 "грн"'
            elif value_type == "number":
                cell.number_format = "0.00"

    total_row = 6 + len(rows)
    label_column = min(
        (
            index
            for index, (key, _, _) in enumerate(selected_columns, start=1)
            if key in totals
        ),
        default=2,
    )
    if label_column > 2:
        sheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=label_column - 1,
        )
    total_label = sheet.cell(row=total_row, column=1, value="Разом")
    total_label.font = Font(bold=True, color="071B3A")
    total_label.alignment = Alignment(horizontal="right")

    for column, (key, _, value_type) in enumerate(
        selected_columns,
        start=1,
    ):
        if key not in totals:
            continue
        cell = sheet.cell(
            row=total_row,
            column=column,
            value=float(totals[key]),
        )
        cell.font = Font(bold=True, color="C85000")
        cell.fill = PatternFill("solid", fgColor="FFF1E8")
        if value_type == "money":
            cell.number_format = '#,##0.00 "грн"'
        elif value_type == "integer":
            cell.number_format = "0"
        else:
            cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")

    for column, (_, header, value_type) in enumerate(
        selected_columns,
        start=1,
    ):
        width = 18
        if value_type == "text":
            width = min(max(len(header) + 4, 20), 42)
        elif value_type == "money":
            width = 18
        sheet.column_dimensions[
            sheet.cell(row=5, column=column).column_letter
        ].width = width
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[5].height = 32

    workbook.save(path)


def export_report_csv(
    path: str,
    period_label: str,
    rows: list[dict[str, object]],
    totals: dict[str, object],
    *,
    report_title: str = "Звіт по замовленнях",
    columns: list[ReportColumn] | None = None,
) -> None:
    selected_columns = columns or REPORT_COLUMNS
    with Path(path).open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow([report_title])
        writer.writerow(["Період", period_label])
        writer.writerow([])
        writer.writerow([header for _, header, _ in selected_columns])
        for row in rows:
            writer.writerow(
                [
                    _csv_value(row.get(key), value_type)
                    for key, _, value_type in selected_columns
                ]
            )
        writer.writerow(
            [
                (
                    _csv_value(totals[key], value_type)
                    if key in totals
                    else ("Разом" if index == 0 else "")
                )
                for index, (key, _, value_type) in enumerate(selected_columns)
            ]
        )


def build_report_html(
    period_label: str,
    rows: list[dict[str, object]],
    totals: dict[str, object],
    *,
    report_title: str = "Звіт по замовленнях",
    columns: list[ReportColumn] | None = None,
) -> str:
    selected_columns = columns or REPORT_COLUMNS
    body_rows = "".join(
        "<tr>"
        + "".join(
            _html_cell(row.get(key), value_type)
            for key, _, value_type in selected_columns
        )
        + "</tr>"
        for row in rows
    )
    headers = "".join(f"<th>{escape(header)}</th>" for _, header, _ in selected_columns)
    footer_cells = "".join(
        (
            f"<td class='{_cell_class(value_type)}'>"
            f"{_display_value(totals[key], value_type)}</td>"
            if key in totals
            else (
                "<td style='text-align:right'>Разом</td>" if index == 0 else "<td></td>"
            )
        )
        for index, (key, _, value_type) in enumerate(selected_columns)
    )
    return f"""
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: "Segoe UI", sans-serif; color: #071B3A; }}
            h1 {{ margin: 0; font-size: 22px; }}
            h2 {{ margin: 5px 0 3px; font-size: 17px; }}
            .meta {{ color: #6F7D95; margin-bottom: 18px; }}
            table {{ width: 100%; border-collapse: collapse; font-size: 8px; }}
            th {{
                background: #06284A; color: white; padding: 7px 5px;
                text-align: left;
            }}
            td {{ border-bottom: 1px solid #DDE5F0; padding: 7px 5px; }}
            .number {{ text-align: right; white-space: nowrap; }}
            .money {{ text-align: right; white-space: nowrap; }}
            tfoot td {{
                background: #FFF1E8; color: #C85000; font-weight: bold;
                border-top: 2px solid #FF6A00;
            }}
        </style>
    </head>
    <body>
        <h1>Creative Spark Agency CRM</h1>
        <h2>{escape(report_title)}</h2>
        <div class="meta">
            Період: {escape(period_label)}<br>
            Сформовано: {datetime.now().strftime("%d.%m.%Y %H:%M")}
        </div>
        <table>
            <thead><tr>{headers}</tr></thead>
            <tbody>{body_rows}</tbody>
            <tfoot><tr>{footer_cells}</tr></tfoot>
        </table>
    </body>
    </html>
    """


def _html_cell(value: object, value_type: str) -> str:
    return (
        f"<td class='{_cell_class(value_type)}'>"
        f"{_display_value(value, value_type)}</td>"
    )


def _cell_class(value_type: str) -> str:
    if value_type == "money":
        return "money"
    if value_type in ("number", "integer"):
        return "number"
    return ""


def _display_value(value: object, value_type: str) -> str:
    if value_type == "money":
        return _money_text(value)
    if value_type == "number":
        return f"{float(value or 0):.2f}"
    if value_type == "integer":
        return str(int(value or 0))
    return escape(str(value or "—"))


def _csv_value(value: object, value_type: str) -> object:
    if value_type in ("money", "number"):
        return _decimal_text(value)
    if value_type == "integer":
        return int(value or 0)
    return value or "—"


def _sheet_title(value: str) -> str:
    cleaned = "".join(
        "_" if character in r"[]:*?/\\" else character for character in value
    )
    return cleaned[:31] or "Звіт"


def _decimal_text(value: object) -> str:
    return f"{float(value or 0):.2f}".replace(".", ",")


def _money_text(value: object) -> str:
    return f"{float(value or 0):,.2f} грн".replace(",", " ")
